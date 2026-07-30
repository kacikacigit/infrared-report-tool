# -*- coding: utf-8 -*-
"""
局放检测超声波图库 - 区域数值提取并写入Excel
======================================================

功能：
1. 交互式框选：打开一张示例图片，用鼠标拖拽框选需要识别数值的区域（ROI），
   该区域坐标会被保存，后续所有图片都使用同一个区域。
2. 批量OCR识别：对文件夹内所有 .jpg 图片，裁剪同样的区域，识别其中的数字
   （可能是负数），并取绝对值。
3. 写入Excel：根据Excel中 E列"图谱文件"的文件名，与图片文件名匹配，
   将识别到的数值（绝对值）写入对应行的 B列。
4. 如果某张图片未能识别出数值，则在B列填写"无法识别"。

依赖库安装（在终端执行）：
    pip install opencv-python pytesseract openpyxl pillow numpy

另外还需要安装 Tesseract OCR 引擎本体（这是识别图片文字的核心程序，
pytesseract 只是它的Python接口，必须单独安装）：
    Mac (使用 Homebrew):
        brew install tesseract
    如果没有安装 Homebrew，先执行:
        /bin/bash -c "$(curl -fsSL https://raw.githubusercontent.com/Homebrew/install/HEAD/install.sh)"

使用方法：
    1. 根据下面 “===== 配置区域 =====” 中的路径确认无误后，直接运行本脚本：
        python3 局放图片数值提取.py
    2. 第一次运行时会弹出一张示例图片窗口，用鼠标左键拖拽框选出你要识别数值
       的区域，框选完成后按【空格键】或【回车键】确认（按 c 键可取消重新框选）。
    3. 框选完成后，程序会自动对文件夹内所有图片进行裁剪+识别，并将结果写入
       Excel文件对应的B列。
    4. 识别区域坐标会保存在 roi_config.json 文件中，下次运行如果该文件存在，
       会询问你是否复用上次的框选区域，无需重新框选。
"""

import os
import re
import json
import glob
import sys

import cv2
import numpy as np
import pytesseract
from openpyxl import load_workbook

# ===================== 配置区域（请根据实际情况核对） =====================

IMAGE_DIR = "/Users/zhanganan/Desktop/人工智能/opencode项目集/局放检测/红外测温图片"
EXCEL_PATH = "/Users/zhanganan/Desktop/人工智能/opencode项目集/局放检测/超声波检测值.xlsx"

# Excel 相关配置
EXCEL_SHEET_NAME = None      # None 表示使用默认的第一个sheet，如需指定sheet名，改成字符串
EXCEL_HEADER_ROW = 1         # 表头所在行号（从1开始计数）
COL_VALUE = "B"              # 数值要写入的列
COL_IMAGE_NAME = "E"         # 图谱文件名所在列

# ROI（识别区域）配置文件，保存框选坐标，方便下次直接复用
ROI_CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "roi_config.json")

# 是否保存裁剪区域及识别结果的调试图片（便于人工核对识别是否准确）
SAVE_DEBUG_CROPS = True
DEBUG_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "debug_crops")

# 无法识别时填写的文字
NOT_RECOGNIZED_TEXT = "无法识别"

# =================================================================


def select_roi_interactively(sample_image_path):
    """打开示例图片，让用户用鼠标框选区域，返回 (x, y, w, h)"""
    img = cv2.imread(sample_image_path)
    if img is None:
        raise FileNotFoundError(f"无法读取示例图片: {sample_image_path}")

    print("\n请在弹出的窗口中用鼠标左键拖拽，框选出需要识别数值的区域。")
    print("框选完成后按【空格键】或【回车键】确认；按 c 键可取消重新框选。\n")

    window_name = "请框选数值区域 (Space/Enter确认, C取消)"
    roi = cv2.selectROI(window_name, img, showCrosshair=True, fromCenter=False)
    cv2.destroyAllWindows()

    x, y, w, h = roi
    if w == 0 or h == 0:
        raise ValueError("未框选有效区域，程序退出。请重新运行脚本。")

    return int(x), int(y), int(w), int(h)


def get_roi(sample_image_path):
    """获取ROI坐标：如果存在配置文件，询问是否复用；否则重新框选。"""
    if os.path.exists(ROI_CONFIG_FILE):
        with open(ROI_CONFIG_FILE, "r", encoding="utf-8") as f:
            cfg = json.load(f)
        x, y, w, h = cfg["x"], cfg["y"], cfg["w"], cfg["h"]
        print(f"检测到已保存的识别区域配置: x={x}, y={y}, w={w}, h={h}")
        choice = input("是否复用该区域？(y=复用 / n=重新框选) [y]: ").strip().lower()
        if choice in ("", "y", "yes"):
            return x, y, w, h

    x, y, w, h = select_roi_interactively(sample_image_path)
    with open(ROI_CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump({"x": x, "y": y, "w": w, "h": h}, f, ensure_ascii=False, indent=2)
    print(f"识别区域已保存到: {ROI_CONFIG_FILE}")
    return x, y, w, h


def preprocess_crop(crop):
    """对裁剪出的小图做预处理，提升OCR识别准确率。"""
    # 放大图片，小字号数字放大后识别效果通常更好
    scale = 4
    crop = cv2.resize(crop, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)

    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)

    # 自动判断是否需要反色：如果背景偏暗（数码管/黑底白字很常见），先反色成黑字白底
    mean_val = np.mean(gray)
    if mean_val < 127:
        gray = cv2.bitwise_not(gray)

    # 去噪 + 二值化（OTSU自动阈值）
    blur = cv2.GaussianBlur(gray, (3, 3), 0)
    _, thresh = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    return thresh


def ocr_extract_abs_value(image_path, roi):
    """裁剪指定区域并识别数字，返回绝对值字符串；识别失败返回 None。"""
    img = cv2.imread(image_path)
    if img is None:
        return None, None

    x, y, w, h = roi
    # 防止越界
    img_h, img_w = img.shape[:2]
    x2, y2 = min(x + w, img_w), min(y + h, img_h)
    x, y = max(x, 0), max(y, 0)
    crop = img[y:y2, x:x2]

    if crop.size == 0:
        return None, None

    processed = preprocess_crop(crop)

    # 只允许识别数字、负号、小数点，psm 7 = 将图像视为单行文本
    custom_config = r'--psm 7 -c tessedit_char_whitelist=0123456789.-'
    text = pytesseract.image_to_string(processed, config=custom_config)

    value = extract_number(text)
    return value, processed


def extract_number(text):
    """从OCR文本中提取数字（可能带负号/小数点），返回绝对值的字符串形式。"""
    if not text:
        return None
    text = text.strip()
    match = re.search(r'-?\d+\.?\d*', text)
    if not match:
        return None
    try:
        num = float(match.group())
        abs_num = abs(num)
        # 如果原本是整数形式，就以整数形式返回，避免多余的 .0
        if abs_num == int(abs_num):
            return str(int(abs_num))
        return str(abs_num)
    except ValueError:
        return None


def collect_image_files(image_dir):
    """获取文件夹内所有 .jpg 图片（不区分大小写）。"""
    patterns = ["*.jpg", "*.JPG", "*.jpeg", "*.JPEG"]
    files = []
    for p in patterns:
        files.extend(glob.glob(os.path.join(image_dir, p)))
    files = sorted(set(files))
    return files


def main():
    if not os.path.isdir(IMAGE_DIR):
        print(f"错误：图片文件夹不存在: {IMAGE_DIR}")
        sys.exit(1)
    if not os.path.isfile(EXCEL_PATH):
        print(f"错误：Excel文件不存在: {EXCEL_PATH}")
        sys.exit(1)

    image_files = collect_image_files(IMAGE_DIR)
    if not image_files:
        print(f"错误：在文件夹中未找到任何 .jpg 图片: {IMAGE_DIR}")
        sys.exit(1)

    print(f"共找到 {len(image_files)} 张图片。")

    # 1. 获取识别区域（用第一张图片作为示例）
    roi = get_roi(image_files[0])

    # 2. 逐张图片识别数值
    if SAVE_DEBUG_CROPS:
        os.makedirs(DEBUG_DIR, exist_ok=True)

    results = {}  # {不带扩展名的文件名: 识别到的值字符串 或 None}
    print("\n开始批量识别，请稍候...\n")
    for i, path in enumerate(image_files, 1):
        base_name = os.path.basename(path)
        name_no_ext = os.path.splitext(base_name)[0]

        value, processed = ocr_extract_abs_value(path, roi)
        results[name_no_ext] = value

        status = value if value is not None else NOT_RECOGNIZED_TEXT
        print(f"[{i}/{len(image_files)}] {base_name}  ->  {status}")

        if SAVE_DEBUG_CROPS and processed is not None:
            debug_name = f"{name_no_ext}_{status}.png"
            debug_name = re.sub(r'[\\/:*?"<>|]', "_", debug_name)  # 文件名安全处理
            cv2.imwrite(os.path.join(DEBUG_DIR, debug_name), processed)

    # 3. 写入Excel
    print("\n正在写入Excel...")
    wb = load_workbook(EXCEL_PATH)
    ws = wb[EXCEL_SHEET_NAME] if EXCEL_SHEET_NAME else wb.active

    matched_count = 0
    unmatched_names = []

    for row in range(EXCEL_HEADER_ROW + 1, ws.max_row + 1):
        cell_e = ws[f"{COL_IMAGE_NAME}{row}"]
        if cell_e.value is None:
            continue
        e_value = str(cell_e.value).strip()
        # 兼容E列写的文件名可能带扩展名，也可能不带
        e_name_no_ext = os.path.splitext(e_value)[0]

        matched_value = None
        if e_value in results:
            matched_value = results[e_value]
            matched_count += 1
        elif e_name_no_ext in results:
            matched_value = results[e_name_no_ext]
            matched_count += 1
        else:
            continue  # 该行图谱文件在图片文件夹中未找到对应图片，跳过

        write_value = matched_value if matched_value is not None else NOT_RECOGNIZED_TEXT
        ws[f"{COL_VALUE}{row}"] = write_value

    wb.save(EXCEL_PATH)

    print(f"\n完成！共匹配并写入 {matched_count} 行数据到: {EXCEL_PATH}")
    if SAVE_DEBUG_CROPS:
        print(f"裁剪区域及识别结果的调试图片已保存到: {DEBUG_DIR}")
        print("如发现识别不准确，可查看调试图片核对，并考虑调整框选区域后重新运行。")


if __name__ == "__main__":
    main()